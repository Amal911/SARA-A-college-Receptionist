from django.shortcuts import render ,redirect
from .models import Event, Attendance, Employees
from .forms import TeacherForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout

# Create your views here.
from django.http import JsonResponse,HttpResponse
import random
import smtplib
from email.mime.text import MIMEText
from . import intents
from django.utils import timezone
from . import recognise, audio, face_train

from openpyxl import Workbook
from datetime import datetime

# Audio
import pyttsx3 as tts
speaker= tts.init()
speaker.setProperty('rate', 150)




def main(request):

    return render(request, 'main.html')




def chatbot(request):
    message = request.GET.get('message')
    print(message)
    if message is not None:
        if 'record' in message:
            cammand = audio()
            response = database(cammand)
            # speaker.say(response)
            # # speaker.runAndWait()
            return JsonResponse(response)

        elif 'mark_attendance' in message:
            reply ={}
            reply['message'] = mark_attendance()
            response=reply
            return JsonResponse(response)
            message = ''
        else:
            response = database(message)
            return JsonResponse(response)
    return render(request,'sara.html')

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('admin')
        else:
            return render(request, 'login1.html', {'error_message': 'Invalid login credentials'})

    return render(request, 'login1.html')

@login_required(login_url='login')
def admin(request):
    if request.method == 'POST':
        today = datetime.now().strftime("%d-%m-%Y")
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Attendance Report %s.xlsx"'%today

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Add column headers
        ws['A1'] = 'Name'
        ws['B1'] = 'Punchin'
        ws['C1'] = 'Punchout'

        # Add data rows
        data = Attendance.objects.values_list('name', 'punchin', 'punchout')
        for row in data:
            ws.append(row)

        wb.save(response)
        return response

        # return HttpResponse("Action performed!")
    return render(request, 'admin_main.html')

@login_required(login_url='login')
def add_teachers(request):
    if request.method == 'POST':
        form = TeacherForm(request.POST, request.FILES)
        if form.is_valid():
            teacher = form.save()

            return render(request, 'employees_dashboard.html', {'teacher': teacher})
    else:
        form = TeacherForm()

    return render(request, 'add_teacher.html',{'form': form})

@login_required(login_url='login')
def add_events(request):
    if request.method == 'POST':
        event_name = request.POST['event_name']
        event_date = request.POST['event_date']
        event = Event(name=event_name, date=event_date)
        event.save()
        return redirect('/event_dashboard')

    return render(request, 'add_event.html')

@login_required(login_url='login')
def event_dashboard(request):

    return render(request, 'event_dashboard1.html')\

@login_required(login_url='login')
def employees_dashboard(request):
    if request.method == 'POST':
        print("done")
        face_train()
    return render(request, 'employees_dashboard.html')



@login_required(login_url='login')
def show_employees(request):
    st = Employees.objects.all()  # Collect all records from table
    return render(request, 'show_employees.html', {'st': st})















def database(message):
    response = {}
    if 'sara' in message.lower():
        message = message.replace('sara','')


    if any( word in message for word in intents.greetings):
        name = recognise()
        if name == "unknown":
            name = ''
        response['message'] = random.choice(intents.greetings_res) +' ' + name

    # Greetings
    elif any( word in message for word in intents.greetings2):
        response['message'] = message

    # Greetings
    elif any( word in message for word in intents.greetingsQ):
        response['message'] = random.choice(intents.greetingsQ_response)

    # Asking Name
    elif any( word in message for word in intents.name):
        response['message'] = random.choice(intents.name_response)

    # Who are you
    elif any( word in message for word in intents.who_are_you):
        response['message'] = random.choice(intents.who_are_you_response)

    # Navigation
    elif any( word in message for word in intents.nav):

        # Principal Office
        if any( word in message for word in intents.nav_principal):
            response['message'] = random.choice(intents.nav_principal_res)

        # Office & Accounts
        elif any(word in message for word in intents.nav_office):
            response['message'] = random.choice(intents.nav_office_res)

        # Admission Cell
        elif any( word in message for word in intents.nav_admission_cell):
            response['message'] = random.choice(intents.nav_admission_cell_res)

        # Library
        elif any( word in message for word in intents.nav_library):
            response['message'] = random.choice(intents.nav_library_res)

        # A Block
        elif any( word in message for word in intents.nav_A_block):
            response['message'] = random.choice(intents.nav_A_block_res)

        # B Block
        elif any(word in message for word in intents.nav_B_block_res):
            response['message'] = random.choice(intents.nav_B_block)

        # Canteen
        elif any(word in message for word in intents.nav_canteen):
            response['message'] = random.choice(intents.nav_canteen_res)

        # Girls Hostel
        elif any(word in message for word in intents.nav_g_hostel):
            response['message'] = random.choice(intents.nav_g_hostel_res)

        #CSE Department
        elif any(word in message for word in intents.nav_cse_dep):
            response['message'] = random.choice(intents.nav_cse_dep_res) + '. To get there, ' +random.choice(intents.nav_A_block_res)

        # EC Department
        elif any(word in message for word in intents.nav_electrical_dep):
            response['message'] = random.choice(intents.nav_electrical_dep_res) + '. To get there, ' + random.choice(intents.nav_A_block_res)

        # EEE Department
        elif any(word in message for word in intents.nav_EEE_dep):
            response['message'] = random.choice(intents.nav_EEE_dep_res) + '. To get there, ' + random.choice(intents.nav_A_block_res)

        # CE Department
        elif any(word in message for word in intents.nav_civil_dep):
            response['message'] = random.choice(intents.nav_civil_dep_res) + '. To get there, ' + random.choice(intents.nav_B_block_res)# CE Department

        # ME Department
        elif any(word in message for word in intents.nav_mech_dep):
            response['message'] = random.choice(intents.nav_mech_dep_res) + '. To get there, ' + random.choice(intents.nav_B_block_res)
        else:
            response['message'] = 'Sorry, I did not understand your message.'

    elif 'set an appointment with principal' in message.lower():
        msg = MIMEText('The user wants to set an appointment with the principal.')
        msg['Subject'] = 'Appointment Request'
        msg['From'] = 'armchatbot3@gmail.com'
        msg['To'] = 'raznalrich@gmail.com'

        # Send the message using SMTP
        smtp_server = 'smtp.gmail.com'
        smtp_port = 587
        smtp_username = 'armchatbot3@gmail.com'
        smtp_password = 'vcefmskrqgtcuvjq'
        smtp_conn = smtplib.SMTP(smtp_server, smtp_port)
        smtp_conn.starttls()
        smtp_conn.login(smtp_username, smtp_password)
        smtp_conn.sendmail(smtp_username, [msg['To']], msg.as_string())
        smtp_conn.quit()

        response['message'] = 'Your appointment request has been sent to the principal.'


    elif 'upcoming events' in message.lower():
        events = Event.objects.all()
        if len(events) > 0:
            event = random.choice(events)
            response['message'] = f"The next event is {event.name} on {event.date}."
        else:
            response['message'] = "There are no upcoming events."

    elif  'want to meet ' in message.lower():
        response['message'] = email(message)

    elif 'mark' in message.lower():
        response['message'] = mark_attendance()



    elif 'time' in message.lower():
        current_time = timezone.now()
        current_time = current_time.strftime("%H:%M")
        print(current_time)
        response['message'] = random.choice(intents.time) + current_time

    elif 'date' in message.lower():
        current_time = timezone.now()
        current_time = current_time.strftime('%d %B %Y')
        response['message'] = random.choice(intents.date) + current_time

    elif 'day' in message.lower():
        current_time = timezone.now()
        current_time = current_time.strftime('%A')
        response['message'] = random.choice(intents.day) + current_time

    elif '@copy' in message.lower():
        Attendance.objects.all().delete()
        people = Employees.objects.all()
        for item in people:
            my_object = Attendance(name=item, punchin='0', punchout='0')
            my_object.save()
        response['message'] = 'done.'

    else:
        response['message'] = 'Sorry, I did not understand your message.'
    return response








def mark_attendance():
    name = recognise()
    if name == 'unknown':
        message = "Sorry You are not Registered"
        return message
    current_time = timezone.now()
    current_time = current_time.strftime("%H:%M:%S")


    if Attendance.objects.filter(name=name).exists() :
        obj = Attendance.objects.get(name=name)
        if obj.punchin != '0':

            obj.punchout = current_time
            obj.save()
            message = "Your punchout is marked. have a nice day " + name
            return message
        elif Attendance.objects.filter(name=name).exists():
            obj = Attendance.objects.get(name=name)
            obj.punchin = current_time
            obj.save()
            message = "Your punchin is marked. have a nice day " + name
            return message
    else:
        message = "Sorry You are not Registered"
        return message



def email(message):

    employee = Employees.objects.all()
    for emp in employee:
        if emp.name.lower() in message.lower():
            name = emp.name
    print(name)

    if Attendance.objects.filter(name=name).exists() :
        obj = Attendance.objects.get(name=name)
        if obj.punchin == '0' or obj.punchout != '0':
            message = "I am sorry, it looks like "+name+" is not available today"
            return message
        else:
            obj = Employees.objects.get(name=name)
            emailid = obj.email

            mail="Greetings "+ name+"\nI wanted to inform you that you have a visitor waiting at the waiting room. Please make sure to attend them quickly\n\nSARA"
            msg = MIMEText(mail)
            msg['Subject'] = 'Visitors Notification'
            msg['From'] = 'armchatbot3@gmail.com'
            msg['To'] = emailid

            # Send the message using SMTP
            smtp_server = 'smtp.gmail.com'
            smtp_port = 587
            smtp_username = 'armchatbot3@gmail.com'
            smtp_password = 'vcefmskrqgtcuvjq'
            smtp_conn = smtplib.SMTP(smtp_server, smtp_port)
            smtp_conn.starttls()
            smtp_conn.login(smtp_username, smtp_password)
            smtp_conn.sendmail(smtp_username, [msg['To']], msg.as_string())
            smtp_conn.quit()

            message="I have informed "+ name + ". Please wait at the waiting room."
            return message